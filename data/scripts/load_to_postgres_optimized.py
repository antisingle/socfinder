#!/usr/bin/env python3
"""
Оптимизированный скрипт для загрузки данных из Excel в PostgreSQL
Читает файл по частям для экономии памяти (для серверов с 1GB RAM)
"""
import os
import sys
import json
import gc
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
import logging

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Добавляем путь к app для импорта моделей
sys.path.append('/app')
from app.models.project import Project, Base

def load_coordinates():
    """Загружаем координаты регионов"""
    try:
        with open('/app/data/regions_coordinates.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Ошибка загрузки координат: {e}")
        return {}

def get_coordinates(region_name, coordinates_dict):
    """Получаем координаты для региона"""
    if not region_name:
        return None
    
    # Ищем точное совпадение
    if region_name in coordinates_dict:
        coords = coordinates_dict[region_name]
        return {"lat": coords["lat"], "lng": coords["lng"]}
    
    # Ищем частичное совпадение
    for region, coords in coordinates_dict.items():
        if region_name.lower() in region.lower() or region.lower() in region_name.lower():
            return {"lat": coords["lat"], "lng": coords["lng"]}
    
    return None

def process_batch(batch_data, session, coordinates_dict, start_row):
    """Обрабатывает батч данных"""
    projects = []
    
    for i, row in enumerate(batch_data):
        row_num = start_row + i
        
        try:
            # Правильное маппинг согласно структуре Excel
            name = str(row[0]) if row[0] else ""
            contest = str(row[1]) if row[1] else ""
            year = int(row[2]) if row[2] and str(row[2]).isdigit() else None
            direction = str(row[3]) if row[3] else ""
            region = str(row[5]) if row[5] else ""  # Колонка 5
            org = str(row[6]) if row[6] else ""     # Колонка 6
            
            # Обрабатываем winner (колонка 11)
            winner = False
            if len(row) > 11 and row[11] is not None:
                winner_val = str(row[11]).lower().strip()
                winner = winner_val in ['true', '1', 'да', 'победитель', 'winner', '+']
            
            # Обрабатываем money_req_grant (колонка 13)
            money_req_grant = 0
            if len(row) > 13 and row[13] is not None:
                try:
                    money_val = row[13]
                    if isinstance(money_val, (int, float)):
                        money_req_grant = int(money_val)
                    else:
                        money_str = str(money_val).replace(' ', '').replace(',', '')
                        if money_str.replace('.', '').isdigit():
                            money_req_grant = int(float(money_str))
                except:
                    money_req_grant = 0
            
            # Создаем проект
            project = Project(
                name=name,
                contest=contest,
                year=year,
                direction=direction,
                region=region,
                org=org,
                winner=winner,
                money_req_grant=money_req_grant,
                coordinates=get_coordinates(region, coordinates_dict)
            )
            
            projects.append(project)
            
        except Exception as e:
            logger.error(f"Ошибка обработки строки {row_num}: {e}")
            continue
    
    # Сохраняем батч
    if projects:
        session.add_all(projects)
        session.commit()
        
    return len(projects)

def main():
    """Основная функция загрузки данных с оптимизацией памяти"""
    logger.info("🚀 Начинаем ОПТИМИЗИРОВАННУЮ загрузку данных в PostgreSQL...")
    logger.info("💾 Режим экономии памяти для серверов с 1GB RAM")
    
    # Подключение к базе данных
    database_url = os.getenv("DATABASE_URL", "postgresql://socfinder:socfinder123@postgres:5432/socfinder")
    logger.info(f"Подключение к базе: {database_url}")
    
    engine = create_engine(database_url, pool_size=1, max_overflow=0)
    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    
    try:
        # Создаем таблицы
        logger.info("Создание таблиц...")
        Base.metadata.create_all(bind=engine)
        
        # Проверяем, есть ли уже данные
        with engine.connect() as conn:
            result = conn.execute(text("SELECT COUNT(*) FROM projects"))
            count = result.scalar()
            if count > 0:
                logger.info(f"В базе уже есть {count} проектов. Пропускаем загрузку.")
                return
        
        # Загружаем координаты
        logger.info("Загрузка координат регионов...")
        coordinates_dict = load_coordinates()
        logger.info(f"Загружено {len(coordinates_dict)} регионов с координатами")
        
        # Проверяем файл
        excel_path = '/app/data/raw/data_114_pres_grants_v20250313.xlsx'
        logger.info(f"Загрузка Excel файла: {excel_path}")
        
        if not os.path.exists(excel_path):
            logger.error(f"Файл не найден: {excel_path}")
            return
        
        # ОПТИМИЗАЦИЯ: Читаем файл с минимальным использованием памяти
        logger.info("📖 Открываем Excel файл в режиме только чтения...")
        workbook = load_workbook(excel_path, read_only=True, data_only=True)
        sheet = workbook.active
        logger.info(f"Открыт лист: {sheet.title}")
        
        # Получаем заголовки (только первая строка)
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        logger.info(f"Заголовки: {headers[:5]}...")
        
        # Создаем сессию
        session = SessionLocal()
        
        try:
            projects_added = 0
            batch_size = 500  # Уменьшенный размер батча для экономии памяти
            batch_data = []
            row_count = 0
            
            logger.info("🔄 Начинаем обработку данных по частям...")
            
            # ОПТИМИЗАЦИЯ: Обрабатываем строки по частям
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_count += 1
                batch_data.append(row)
                
                # Обрабатываем батч когда он заполнен
                if len(batch_data) >= batch_size:
                    # Обрабатываем батч
                    added = process_batch(batch_data, session, coordinates_dict, row_count - len(batch_data) + 1)
                    projects_added += added
                    
                    # КРИТИЧНО: Очищаем память
                    batch_data = []
                    gc.collect()  # Принудительная сборка мусора
                    
                    logger.info(f"✅ Сохранено проектов: {projects_added} (обработано строк: {row_count})")
                    
                    # Небольшая пауза для системы
                    import time
                    time.sleep(0.1)
                
                # Прогресс каждые 5000 строк
                if row_count % 5000 == 0:
                    logger.info(f"📊 Обработано строк: {row_count}")
            
            # Обрабатываем оставшиеся данные
            if batch_data:
                added = process_batch(batch_data, session, coordinates_dict, row_count - len(batch_data) + 1)
                projects_added += added
                batch_data = []
                gc.collect()
            
            logger.info(f"🎉 Загрузка завершена!")
            logger.info(f"📊 Всего строк обработано: {row_count}")
            logger.info(f"✅ Всего проектов добавлено: {projects_added}")
            
        except Exception as e:
            logger.error(f"❌ Ошибка при загрузке: {e}")
            session.rollback()
            raise
        finally:
            session.close()
            workbook.close()  # Освобождаем память
            gc.collect()
            
    except Exception as e:
        logger.error(f"💥 Критическая ошибка: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
