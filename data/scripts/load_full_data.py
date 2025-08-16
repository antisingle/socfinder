#!/usr/bin/env python3
"""
Скрипт для загрузки ВСЕХ данных из Excel в SQLite базу данных
"""
import sqlite3
import json
import os
from openpyxl import load_workbook
from datetime import datetime

def load_regions_coordinates():
    """Загружает координаты регионов"""
    coords_path = '/app/data/regions_coordinates.json'
    with open(coords_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def clean_value(value):
    """Очищает значение от пустых строк"""
    if value is None or value == '':
        return None
    return value

def parse_date(date_str):
    """Парсит дату из строки"""
    if not date_str or date_str == '':
        return None
    
    # Если это уже datetime объект
    if hasattr(date_str, 'strftime'):
        return date_str.strftime('%Y-%m-%d %H:%M:%S')
    
    # Пробуем разные форматы дат
    date_formats = [
        '%Y-%m-%d',
        '%d.%m.%Y',
        '%d/%m/%Y',
        '%Y-%m-%d %H:%M:%S'
    ]
    
    for fmt in date_formats:
        try:
            dt = datetime.strptime(str(date_str), fmt)
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            continue
    
    return None

def parse_boolean(value):
    """Парсит булево значение"""
    if value is None or value == '':
        return None
    
    value_str = str(value).lower().strip()
    if value_str in ['да', 'yes', 'true', '1']:
        return True
    elif value_str in ['нет', 'no', 'false', '0']:
        return False
    
    return None

def parse_number(value):
    """Парсит числовое значение"""
    if value is None or value == '':
        return None
    
    try:
        # Убираем пробелы и заменяем запятые на точки
        value_str = str(value).replace(' ', '').replace(',', '.')
        return float(value_str)
    except (ValueError, TypeError):
        return None

def create_database():
    """Создает базу данных и таблицы"""
    print("Создание базы данных...")
    
    conn = sqlite3.connect('/app/socfinder.db')
    cursor = conn.cursor()
    
    # Удаляем старую таблицу если есть
    cursor.execute('DROP TABLE IF EXISTS projects')
    
    # Создаем таблицу projects
    cursor.execute('''
    CREATE TABLE projects (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        contest TEXT,
        year INTEGER,
        direction TEXT,
        date_req DATETIME,
        region TEXT,
        org TEXT,
        inn TEXT,
        ogrn TEXT,
        implem_start DATETIME,
        implem_end DATETIME,
        winner BOOLEAN,
        rate REAL,
        money_req_grant INTEGER,
        cofunding INTEGER,
        total_money INTEGER,
        description TEXT,
        goal TEXT,
        tasks TEXT,
        soc_signif TEXT,
        pj_geo TEXT,
        target_groups TEXT,
        address TEXT,
        web_site TEXT,
        req_num TEXT,
        link TEXT,
        okato TEXT,
        oktmo TEXT,
        level TEXT,
        coordinates TEXT
    )
    ''')
    
    # Создаем индексы для быстрого поиска
    indexes = [
        'CREATE INDEX idx_region ON projects(region)',
        'CREATE INDEX idx_year ON projects(year)',
        'CREATE INDEX idx_direction ON projects(direction)',
        'CREATE INDEX idx_winner ON projects(winner)',
        'CREATE INDEX idx_org ON projects(org)',
        'CREATE INDEX idx_name ON projects(name)'
    ]
    
    for index in indexes:
        cursor.execute(index)
    
    conn.commit()
    conn.close()
    print("База данных создана")

def load_excel_to_db():
    """Основная функция загрузки данных"""
    print("Начинаем загрузку ВСЕХ данных из Excel...")
    
    # Проверяем наличие файла
    excel_file = '/app/data/raw/data_114_pres_grants_v20250313.xlsx'
    if not os.path.exists(excel_file):
        print(f"ОШИБКА: Файл {excel_file} не найден!")
        return False
    
    try:
        # Загружаем координаты регионов
        coordinates_map = load_regions_coordinates()
        
        # Создаем базу данных
        create_database()
        
        # Читаем Excel файл
        print("Загрузка Excel файла...")
        workbook = load_workbook(excel_file, read_only=True)
        sheet = workbook.active
        
        # Получаем заголовки из первой строки
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        print(f"Найдено колонок: {len(headers)}")
        print(f"Заголовки: {headers[:10]}...")  # Показываем первые 10
        
        # Подключаемся к базе данных
        conn = sqlite3.connect('/app/socfinder.db')
        cursor = conn.cursor()
        
        # Обрабатываем данные построчно
        row_count = 0
        inserted_count = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row_idx % 1000 == 0:
                print(f"Обработано строк: {row_idx}")
            
            row_count += 1
            
            # Пропускаем пустые строки
            if not any(row):
                continue
            
            try:
                # Маппим данные по позициям (адаптируем под реальную структуу Excel)
                # Основные поля
                name = clean_value(row[0]) if len(row) > 0 else None
                contest = clean_value(row[1]) if len(row) > 1 else None
                year = parse_number(row[2]) if len(row) > 2 else None
                direction = clean_value(row[3]) if len(row) > 3 else None
                date_req = parse_date(row[4]) if len(row) > 4 else None
                region = clean_value(row[5]) if len(row) > 5 else None
                org = clean_value(row[6]) if len(row) > 6 else None
                inn = clean_value(row[7]) if len(row) > 7 else None
                ogrn = clean_value(row[8]) if len(row) > 8 else None
                implem_start = parse_date(row[9]) if len(row) > 9 else None
                implem_end = parse_date(row[10]) if len(row) > 10 else None
                winner = parse_boolean(row[11]) if len(row) > 11 else None
                rate = parse_number(row[12]) if len(row) > 12 else None
                money_req_grant = parse_number(row[13]) if len(row) > 13 else None
                cofunding = parse_number(row[14]) if len(row) > 14 else None
                total_money = parse_number(row[15]) if len(row) > 15 else None
                description = clean_value(row[16]) if len(row) > 16 else None
                goal = clean_value(row[17]) if len(row) > 17 else None
                tasks = clean_value(row[18]) if len(row) > 18 else None
                soc_signif = clean_value(row[19]) if len(row) > 19 else None
                pj_geo = clean_value(row[20]) if len(row) > 20 else None
                target_groups = clean_value(row[21]) if len(row) > 21 else None
                address = clean_value(row[22]) if len(row) > 22 else None
                web_site = clean_value(row[23]) if len(row) > 23 else None
                req_num = clean_value(row[24]) if len(row) > 24 else None
                link = clean_value(row[25]) if len(row) > 25 else None
                okato = clean_value(row[26]) if len(row) > 26 else None
                oktmo = clean_value(row[27]) if len(row) > 27 else None
                level = clean_value(row[28]) if len(row) > 28 else None
                
                # Добавляем координаты
                coordinates = None
                if region and region in coordinates_map:
                    coordinates = json.dumps(coordinates_map[region])
                elif region:
                    # Москва по умолчанию для неизвестных регионов
                    coordinates = json.dumps({"lat": 55.7558, "lng": 37.6173})
                
                # Вставляем в базу
                cursor.execute('''
                INSERT INTO projects (
                    name, contest, year, direction, date_req, region, org, inn, ogrn,
                    implem_start, implem_end, winner, rate, money_req_grant, cofunding,
                    total_money, description, goal, tasks, soc_signif, pj_geo,
                    target_groups, address, web_site, req_num, link, okato, oktmo, level,
                    coordinates
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    name, contest, int(year) if year else None, direction, date_req, region, org, inn, ogrn,
                    implem_start, implem_end, winner, rate, 
                    int(money_req_grant) if money_req_grant else None,
                    int(cofunding) if cofunding else None,
                    int(total_money) if total_money else None,
                    description, goal, tasks, soc_signif, pj_geo,
                    target_groups, address, web_site, req_num, link, okato, oktmo, level,
                    coordinates
                ))
                
                inserted_count += 1
                
            except Exception as e:
                print(f"Ошибка в строке {row_idx}: {e}")
                continue
        
        conn.commit()
        conn.close()
        workbook.close()
        
        print(f"✅ Данные успешно загружены!")
        print(f"Всего строк обработано: {row_count}")
        print(f"Проектов добавлено в базу: {inserted_count}")
        return True
        
    except Exception as e:
        print(f"❌ ОШИБКА при загрузке данных: {e}")
        return False

if __name__ == "__main__":
    success = load_excel_to_db()
    exit(0 if success else 1)


